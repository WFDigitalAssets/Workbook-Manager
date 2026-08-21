"""
sql_reconcile.py
----------------
Checks each workbook tag against the historian (ROC-HIST01) to see whether the
tag actually has data being written to it, then produces:

  1. A full "Status" column (column A) that matches the workbook ROW-FOR-ROW,
     so it can be pasted straight into A1 of the sheet. Header rows emit
     "Status", blank/border rows emit "" (empty, so pasting leaves the black
     border formatting untouched), data rows emit the reconciled status.

  2. A list of DISCREPANCIES for the user to reconcile manually.

Reconciliation rules (per tag):
  - Status blank (NEW) + writing in SQL   -> FLAG (new tag already has data)
  - Status "In Progress"  + writing       -> FLAG (in progress but data present)
  - Status "Faulty"       + writing       -> FLAG (faulty but data present)
  - Status "Complete"     + not writing    -> FLAG (complete but no data)
  - status is always taken from the workbook; SQL only raises flags, it never
    changes which bucket a tag is in.

Only rows that are NEW, In Progress, Faulty, or Complete are checked against
SQL. Untouched rows (blank status, PLC fields not filled) are ignored.

NOTE: is_writing_in_sql() is a STUB until the historian schema is confirmed.
Everything else is complete and testable with a fake writer.
"""

import os
import openpyxl
from workbook_manager import (
    _norm, _detect_columns, is_new_tag, _load_workbook_safe,
    STATUS_IN_PROGRESS, STATUS_FAULTY, STATUS_COMPLETE,
)

# Row-type constants
ROW_HEADER = "header"
ROW_BORDER = "border"   # blank or black-fill separator
ROW_DATA   = "data"


def classify_rows(path):
    """
    Walk the sheet and return a list of (row_number, row_type, data_dict|None),
    one entry per physical row, in order. This preserves the exact vertical
    layout so the output column lines up on paste-back.

    Column positions are detected per table from the header row, so mills with
    different layouts (e.g. Armour has no Status column) parse correctly.
    """
    wb, _tmp = _load_workbook_safe(path)
    ws = wb.worksheets[0]        # always the first sheet, not the last-active one
    ncols = ws.max_column
    out = []
    cols = None
    for r in range(1, ws.max_row + 1):
        header = _detect_columns(ws, r, ncols)
        if header is not None:
            cols = header
            out.append((r, ROW_HEADER, None))
            continue

        def get(key):
            c = cols.get(key) if cols else None
            return _norm(ws.cell(row=r, column=c).value) if c else ""

        table = get("table")
        tag   = get("tag_name")
        if not table and not tag:
            out.append((r, ROW_BORDER, None))   # blank or black-fill row
            continue
        data = {
            "status":     get("status"),
            "table":      table,
            "tag_name":   tag,
            "plc_tag":    get("plc_tag"),
            "plc_name":   get("plc_name"),
            "plc_ip":     get("plc_ip"),
            "row_number": r,
        }
        out.append((r, ROW_DATA, data))

    if _tmp and os.path.exists(_tmp):
        try:
            os.remove(_tmp)
        except OSError:
            pass
    return out


# ======================================================================
#  SQL check — queries _PlcOptTagStatus_AllMills once, caches, looks up.
# ======================================================================
# pyodbc is imported lazily inside the functions that need a live connection,
# so this module (and the fake-writer tests) load fine without an ODBC driver.

# Mills present in the central rollup (have a HIST_ linked server).
# A tag in any mill NOT in this set can't be determined -> we don't flag it.
CONNECTED_MILLS = {"ANG", "ARM", "DGM", "HNM", "JOY",
                   "LEO", "MAP", "NEW", "NWB", "OPE"}

# System / metadata columns that aren't real user tags. The historian view
# leaks some of these (known gap in _UpdatePlcOptTagStatus), and the workbooks
# also list them as rows — so we ignore them on BOTH sides: they're never
# treated as "writing" and never reconciled. Compared upper-cased.
SYSTEM_COLUMNS = {
    "DATETIMESTAMP", "RECORDID", "DATETIMEWRITTEN",
    "TMSTAMP", "DATETIMESTAMPWRITTEN",
}


def load_writing_set(sql_conn):
    """
    One round trip: pull every actively-writing PLC column across all mills
    from dbo._PlcOptTagStatus_AllMills. Returns a set of
    (site_upper, table_upper, column_upper) for O(1) lookup.

    Only schema 'dbo' (PLC) and Status 'active' rows are included — that's
    exactly "this PLC tag is writing data."
    """
    query = """
        SELECT Site, [Table], [Column]
        FROM dbo._PlcOptTagStatus_AllMills
        WHERE [Schema] = 'dbo' AND Status = 'active';
    """
    writing = set()
    cursor = sql_conn.cursor()
    # The view UNION ALLs across all 10 mills' linked servers, so it can be
    # slow (esp. newer/farther mills like OPE). Give it plenty of time; 0 = no
    # limit. Raise the connection's query timeout for this heavy query.
    sql_conn.timeout = 300          # seconds for query execution
    cursor.execute(query)
    for site, table, column in cursor.fetchall():
        col_upper = _norm(column).upper()
        if col_upper in SYSTEM_COLUMNS:
            continue                     # ignore metadata/system columns
        writing.add((
            _norm(site).upper(),
            _norm(table).upper(),
            col_upper,
        ))
    return writing


def make_writer(sql_conn):
    """
    Build a writer function closed over a single cached query result.
    Returns (writer, connected_check) where:
      writer(conn, mill, table, tag) -> True/False   (is it writing)
      Only call writer for mills in CONNECTED_MILLS; others are 'unknown'.
    """
    writing_set = load_writing_set(sql_conn)

    def writer(_conn, mill_code, table_name, tag_name):
        key = (mill_code.upper(), table_name.upper(), tag_name.upper())
        return key in writing_set

    return writer


def is_writing_in_sql(sql_conn, mill_code, table_name, tag_name):
    """
    Convenience single-tag check (builds the cache each call — inefficient).
    Prefer make_writer() when checking many tags. Kept for compatibility.
    """
    writer = make_writer(sql_conn)
    return writer(sql_conn, mill_code, table_name, tag_name)


# ======================================================================
#  Reconciliation
# ======================================================================
def reconcile_row(data, writing):
    """
    Given a data row dict and whether SQL says it's writing, return
    (new_status, flag_or_None).
      new_status : the value to put in column A for this row
      flag       : a human-readable discrepancy message, or None
    """
    status = data["status"].lower()

    # NEW = blank status but all three PLC fields filled.
    if status == "" and is_new_tag(data):
        if writing:
            return "", (f"NEW tag already writing in SQL: "
                        f"{data['table']}.{data['tag_name']}")
        return "", None

    if status == STATUS_IN_PROGRESS:
        if writing:
            return data["status"], (f"IN PROGRESS tag is writing in SQL: "
                                    f"{data['table']}.{data['tag_name']}")
        return data["status"], None

    if status == STATUS_FAULTY:
        if writing:
            return data["status"], (f"FAULTY tag is writing in SQL: "
                                    f"{data['table']}.{data['tag_name']}")
        return data["status"], None

    if status == STATUS_COMPLETE:
        if not writing:
            return data["status"], (f"COMPLETE tag has no data in SQL: "
                                    f"{data['table']}.{data['tag_name']}")
        return data["status"], None

    # Any other / untouched row: unchanged, not flagged.
    return data["status"], None


def needs_sql_check(data):
    """Only NEW / In Progress / Faulty / Complete rows are checked.
    System/metadata columns (RecordID, TmStamp, etc.) are never checked."""
    if data["tag_name"].upper() in SYSTEM_COLUMNS:
        return False
    s = data["status"].lower()
    if s in (STATUS_IN_PROGRESS, STATUS_FAULTY, STATUS_COMPLETE):
        return True
    return s == "" and is_new_tag(data)


def build_status_column(path, sql_conn, mill_code, writer=None):
    """
    Produce:
      column : list[str] of length == sheet rows, in order (paste into A1)
      flags  : list[str] of discrepancy messages for manual reconciliation

    If the mill isn't in the central rollup (CONNECTED_MILLS), SQL status is
    UNKNOWN: statuses are left exactly as-is and nothing is flagged, since we
    can't tell whether anything is writing.

    `writer` is injectable for testing. If None and mill is connected, a real
    writer is built from `sql_conn` (one cached query).
    """
    rows = classify_rows(path)

    connected = mill_code.upper() in CONNECTED_MILLS
    if writer is None and connected:
        writer = make_writer(sql_conn)

    column = []
    flags = []
    for row_num, row_type, data in rows:
        if row_type == ROW_HEADER:
            column.append("Status")
        elif row_type == ROW_BORDER:
            column.append("")            # empty -> leaves black border intact
        else:  # DATA
            if not connected:
                # Can't determine writing status; leave untouched, no flags.
                column.append(data["status"])
                continue
            if needs_sql_check(data):
                writing = writer(sql_conn, mill_code, data["table"], data["tag_name"])
            else:
                writing = False
            new_status, flag = reconcile_row(data, writing)
            column.append(new_status)
            if flag:
                flags.append(f"row {row_num}: {flag}")

    if not connected:
        flags.append(f"NOTE: {mill_code} is not connected to the historian rollup "
                     f"— statuses left unchanged, no SQL check performed.")
    return column, flags


# ======================================================================
#  Excel output — one sheet, all mills' column A side by side.
# ======================================================================
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# "Black, Text 1" theme color -> pure black solid fill for border rows.
BORDER_FILL = PatternFill(start_color="FF000000", end_color="FF000000",
                          fill_type="solid")


def build_reconcile_workbook(folder, filename_to_mill, sql_conn, out_path):
    """
    For every workbook in `folder`, reconcile against SQL and write ONE output
    .xlsx:
      - Sheet "Statuses": each mill = one column, containing that workbook's
        full column A (headers say 'Status', border rows painted black,
        data rows = reconciled status). Copy a column, paste into the real
        workbook's column A.
      - Sheet "Flags": all manual-reconciliation items, per mill.
    """
    import glob, os

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "Statuses"
    flags_ws = wb_out.create_sheet("Flags")

    files = sorted(f for f in glob.glob(os.path.join(folder, "*.xlsx"))
                   if not os.path.basename(f).startswith("~$")
                   and os.path.basename(f) != "reconcile_output.xlsx")

    all_flags = []      # (mill, message)
    col_idx = 1
    mill_order = []
    for path in files:
        name = os.path.basename(path)
        mill = filename_to_mill.get(name, os.path.splitext(name)[0])
        mill_order.append(mill)

        column, flags = build_status_column(path, sql_conn, mill)

        # Output column starts at ROW 1 to line up exactly with the workbook's
        # column A (row 1 = 'Status' header, etc.). No offset — paste into A1.
        row_types = [rt for (_, rt, _) in classify_rows(path)]

        # Determine which output rows get the black border: the row directly
        # ABOVE each header after the first. That's the reliable position of
        # the black separator, rather than guessing from blank/border rows.
        header_positions = [i for i, rt in enumerate(row_types)
                            if rt == ROW_HEADER]
        border_positions = {h - 1 for h in header_positions[1:] if h - 1 >= 0}

        for i, value in enumerate(column):
            out_cell = ws.cell(row=i + 1, column=col_idx, value=value)
            if i in border_positions:
                out_cell.fill = BORDER_FILL

        for f in flags:
            all_flags.append((mill, f))
        col_idx += 1

    # Flags sheet — start with the mill -> column mapping so you know which
    # output column is which mill (the status columns can't carry a header
    # without breaking row alignment).
    flags_ws.cell(row=1, column=1, value="Column").font = Font(bold=True)
    flags_ws.cell(row=1, column=2, value="Mill").font = Font(bold=True)
    from openpyxl.utils import get_column_letter
    for i, mill in enumerate(mill_order):
        flags_ws.cell(row=i + 2, column=1, value=get_column_letter(i + 1))
        flags_ws.cell(row=i + 2, column=2, value=mill)

    # then the discrepancies below, with a gap
    start = len(mill_order) + 3
    flags_ws.cell(row=start, column=1, value="Mill").font = Font(bold=True)
    flags_ws.cell(row=start, column=2, value="Discrepancy").font = Font(bold=True)
    for r, (mill, msg) in enumerate(all_flags, start=start + 1):
        flags_ws.cell(row=r, column=1, value=mill)
        flags_ws.cell(row=r, column=2, value=msg)

    wb_out.save(out_path)
    return out_path, len(all_flags), mill_order