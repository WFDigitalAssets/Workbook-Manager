"""
workbook_manager.py
-------------------
Manages the PLC tag workbooks sent out to the mills.

Two options for now:
  1. Submit to Tag Request Form  -> produces a copy-ready table of NEW tags,
                                     mapped into the Tag Request Form columns.
  2. Generate report             -> per-mill counts of New / In Progress /
                                     Faulty / Complete, with a Comments column.

Workbook structure (confirmed against real files):
  - One mill per file, single sheet.
  - The sheet holds several tables stacked vertically.
  - Each table starts with its own header row:
        Status | TableName | StandardTagName | PLC Tag | PLC Name |
        PLC IP Address | DataType | TagPurpose | DefaultSource | Notes | Criticality
  - Tables are separated by a blank row and a black-filled row. Those are
    unreadable by fill color via openpyxl, so we DON'T rely on them. Instead we
    treat every header row as the start of a new table and skip any row that
    isn't a real data row.

Definitions:
  - A row is a DATA row if it has a TableName and StandardTagName.
  - PLC Tag == "N/A" counts as NOT filled (ignore it).
  - A NEW tag = PLC Tag filled (and != N/A) AND PLC Name filled AND
    PLC IP Address filled AND Status is blank.
"""

import os
import glob
import subprocess
import tempfile
import openpyxl


RECONCILE_OUTPUT_NAME = "reconcile_output.xlsx"

def list_workbook_files(folder):
    """All input workbooks in the folder, excluding Excel lock files and the
    reconcile output file (so it never processes its own output)."""
    return sorted(
        f for f in glob.glob(os.path.join(folder, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
        and os.path.basename(f) != RECONCILE_OUTPUT_NAME
    )


def _get_sql_connection():
    """
    Open a pyodbc connection to ROC-HIST01 using credentials from a .env file
    (never hardcoded). Expects these keys in .env:
        HIST_SERVER=ROC-HIST01
        HIST_DATABASE=Historian
        HIST_UID=Reyadmin
        HIST_PWD=your_password_here
    pyodbc and python-dotenv are imported lazily so the rest of the program
    runs without them.
    """
    import pyodbc
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass  # if python-dotenv isn't installed, rely on real env vars

    server = os.getenv("HIST_SERVER", "ROC-HIST01")
    database = os.getenv("HIST_DATABASE", "Historian")
    uid = os.getenv("HIST_UID", "")
    pwd = os.getenv("HIST_PWD", "")
    if not uid or not pwd:
        raise RuntimeError(
            "Missing HIST_UID / HIST_PWD. Put them in a .env file "
            "(and make sure .env is in .gitignore)."
        )

    # Password braced so special characters don't break the ODBC string.
    conn_str = (
        "Driver={ODBC Driver 18 for SQL Server};"
        f"Server={server};"
        f"Database={database};"
        f"UID={uid};"
        "PWD={" + pwd + "};"
        "Encrypt=no;"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str, timeout=10)

# ----- Column positions (1-indexed), fixed by the workbook template -----
COL_STATUS      = 1
COL_TABLENAME   = 2
COL_TAGNAME     = 3
COL_PLCTAG      = 4
COL_PLCNAME     = 5
COL_PLCIP       = 6
COL_DATATYPE    = 7
COL_TAGPURPOSE  = 8

# The exact text of a header row's first columns, used to detect table starts.
HEADER_SIGNATURE = ("Status", "TableName", "StandardTagName")

# Status values (compared case-insensitively).
STATUS_IN_PROGRESS = "in progress"
STATUS_FAULTY      = "faulty"
STATUS_COMPLETE    = "complete"

# Mills expected in the report, in display order.
REPORT_MILLS = ["ANG", "ARM", "DGM", "HNM", "JOY", "MAP", "NEW", "NWB", "RUS"]

# Explicit filename -> mill code map (filenames as sent out to the mills).
FILENAME_TO_MILL = {
    "Angelina PLC Tags.xlsx":     "ANG",
    "Armour PLC Tags.xlsx":       "ARM",
    "Dudley PLC Tags.xlsx":       "DGM",
    "Henderson PLC Tags.xlsx":    "HNM",
    "Joyce PLC Tags.xlsx":        "JOY",
    "Maplesville PLC Tags.xlsx":  "MAP",
    "New Boston PLC Tags.xlsx":   "NEW",
    "Newberry PLC Tags.xlsx":     "NWB",
    "Opelika PLC Tags.xlsx":      "OPE",
    "Russellville PLC Tags.xlsx": "RUS",
}

def mill_code_from_filename(path):
    """Look up the mill code by exact filename; fall back to the stem if unknown."""
    name = os.path.basename(path)
    if name in FILENAME_TO_MILL:
        return FILENAME_TO_MILL[name]
    return os.path.splitext(name)[0]


def _norm(value):
    """Trim to a clean string; None and whitespace-only (incl. non-breaking
    spaces \\xa0) become ''."""
    if value is None:
        return ""
    s = str(value).replace("\xa0", " ").strip()
    return s


def _is_header_row(ws, row, ncols):
    """A header row is one that contains the TableName and StandardTagName
    signature somewhere in it. Matched flexibly (startswith, case-insensitive)
    so variants like 'StandardTagName DataParc' still count."""
    return _detect_columns(ws, row, ncols) is not None


# Header label -> the field key we store it under. Matched by "does the cell
# text START WITH this label" (case-insensitive), so 'PLC Tag Control Logix'
# matches 'plc tag', 'StandardTagName DataParc' matches 'standardtagname', etc.
_HEADER_MAP = [
    ("status",          "status"),
    ("tablename",       "table"),
    ("standardtagname", "tag_name"),
    ("plc tag",         "plc_tag"),
    ("plc name",        "plc_name"),
    ("plc ip",          "plc_ip"),
    ("datatype",        "data_type"),
    ("tagpurpose",      "tag_purpose"),
    ("defaultsource",   "default_source"),
    ("notes",           "notes"),
    ("criticality",     "criticality"),
]


def _detect_columns(ws, row, ncols):
    """
    Look at `row` and, if it's a header row, return a dict mapping each field
    key -> its 1-indexed column. Returns None if this isn't a header row.

    A row qualifies as a header if it contains BOTH a TableName column and a
    StandardTagName column (the minimal signature every mill shares). Mills
    that lack a Status column (e.g. Armour) simply won't have 'status' in the
    returned map — callers treat a missing status as blank.
    """
    found = {}
    for c in range(1, ncols + 1):
        cell = _norm(ws.cell(row=row, column=c).value).lower()
        if not cell:
            continue
        for label, key in _HEADER_MAP:
            if key in found:
                continue                 # first match wins for each field
            if cell.startswith(label):
                found[key] = c
                break
    # Must have at least TableName + StandardTagName to be a real header.
    if "table" in found and "tag_name" in found:
        return found
    return None


def _plc_tag_filled(plc_tag):
    """PLC Tag is 'filled' only if non-empty and not the literal N/A placeholder."""
    return plc_tag != "" and plc_tag.upper() != "N/A"


def count_tags(path):
    """
    Total tags in a workbook = rows where BOTH the TableName and the
    StandardTagName columns have content. Header rows and blank/separator rows
    are already excluded by read_rows; this additionally drops any row missing
    either of the two required columns. This is the authoritative 'total tags'
    count for a mill.
    """
    return sum(
        1 for r in read_rows(path)
        if r["table"].strip() and r["tag_name"].strip()
    )


def _load_workbook_safe(path):
    """
    Open a workbook for reading WITHOUT being blocked by an Excel lock.

    If the file is open in Excel, openpyxl can fail or hang on the lock. To avoid
    that (important for the unattended scheduled run), copy the file to a temp
    location first and parse the copy — a read-copy succeeds even while the
    original is open. Returns (workbook, temp_path); temp_path is None if we fell
    back to opening the original. Caller must delete temp_path when done.
    """
    import shutil
    tmp = None
    try:
        fd, tmp = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        shutil.copy2(path, tmp)
        return openpyxl.load_workbook(tmp, data_only=True), tmp
    except Exception:
        if tmp and os.path.exists(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass
        # Best-effort fallback: open the original directly.
        return openpyxl.load_workbook(path, data_only=True), None


def read_rows(path):
    """
    Read one workbook and yield a dict per DATA row.

    Column positions are detected from each table's header row (not assumed
    fixed), so mills with different layouts all parse correctly:
      - Armour has no Status column -> status defaults to "".
      - Opelika's headers read 'StandardTagName DataParc' etc. -> matched by
        prefix.
    Header rows and blank/separator rows are skipped automatically.

    Uses a copy-first open so a workbook left open in Excel can't lock/hang the
    parse (matters for the scheduled run).
    """
    wb, _tmp = _load_workbook_safe(path)
    # Always use the FIRST sheet. wb.active returns whichever sheet was selected
    # when the file was last saved, which on multi-sheet workbooks can be the
    # wrong one — so index explicitly.
    ws = wb.worksheets[0]
    ncols = ws.max_column
    rows = []
    cols = None                          # current table's column map

    for r in range(1, ws.max_row + 1):
        header = _detect_columns(ws, r, ncols)
        if header is not None:
            cols = header                # new table starts here
            continue
        if cols is None:
            continue                     # data before any header -> skip

        def get(key):
            c = cols.get(key)
            return _norm(ws.cell(row=r, column=c).value) if c else ""

        table = get("table")
        tag   = get("tag_name")
        # Not a data row (blank line, black separator, stray note) -> skip.
        if not table and not tag:
            continue

        rows.append({
            "status":         get("status"),      # "" if this mill has no Status col
            "table":          table,
            "tag_name":       tag,
            "plc_tag":        get("plc_tag"),
            "plc_name":       get("plc_name"),
            "plc_ip":         get("plc_ip"),
            "data_type":      get("data_type"),
            "tag_purpose":    get("tag_purpose"),
            "default_source": get("default_source"),
            "notes":          get("notes"),
            "criticality":    get("criticality"),
            "row_number":     r,
        })

    # Clean up the temp copy (if we made one).
    if _tmp and os.path.exists(_tmp):
        try:
            os.remove(_tmp)
        except OSError:
            pass
    return rows


def is_new_tag(row):
    """New = three PLC inputs filled (PLC Tag != N/A) and Status blank."""
    return (
        _plc_tag_filled(row["plc_tag"])
        and row["plc_name"] != ""
        and row["plc_ip"] != ""
        and row["status"] == ""
    )


# ======================================================================
#  OPTION 1 — Submit to Tag Request Form
# ======================================================================
# Tag Request Form columns:
#   Table Name | Tag name | plc ip address | plc path | data type |
#   units | frequency | description | min | max | newtable?
#
# Mapping from workbook -> form:
#   TableName              -> Table Name
#   StandardTagName        -> Tag name
#   PLC Name + " " + PLC IP -> plc ip address
#   PLC Tag                -> plc path
#   DataType               -> data type
#   (units, frequency, description, min, max, newtable? left blank for now)

FORM_HEADERS = [
    "Table Name", "Tag name", "plc ip address", "plc path", "data type",
    "units", "frequency", "description", "min", "max", "newtable?",
]

def build_submit_table(path):
    """Return (header, rows) for the copy-ready Tag Request Form table."""
    out = []
    for row in read_rows(path):
        if not is_new_tag(row):
            continue
        out.append([
            row["table"],                              # Table Name
            row["tag_name"],                           # Tag name
            f'{row["plc_name"]} {row["plc_ip"]}',      # plc ip address (combined)
            row["plc_tag"],                            # plc path
            row["data_type"],                          # data type
            "",                                        # units
            "",                                        # frequency
            row["tag_purpose"],                        # description <- TagPurpose
            "",                                        # min
            "",                                        # max
            "",                                        # newtable?
        ])
    return FORM_HEADERS, out


def print_submit_table(path):
    header, rows = build_submit_table(path)
    mill = mill_code_from_filename(path)
    if not rows:
        print(f"\n[{mill}] No new tags to submit "
              f"(no rows with all 3 PLC fields filled and a blank status).\n")
        return
    # Build tab-separated text (header + rows).
    lines = ["\t".join(header)]
    for r in rows:
        lines.append("\t".join(r))
    text = "\n".join(lines)

    # Write to a temp .txt file and open it in Notepad, ready to copy/paste.
    # Tabs are preserved in the file, so pasting into the form works correctly.
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=f"_{mill}_submit.txt", delete=False, encoding="utf-8")
    tmp.write(text)
    tmp.close()

    print(f"\n[{mill}] {len(rows)} new tag(s) ready — opening in Notepad.")
    print(f"Select all (Ctrl+A), copy (Ctrl+C), paste into the form.\n")
    try:
        os.startfile(tmp.name)          # Windows: opens in default .txt app (Notepad)
    except AttributeError:
        # Non-Windows fallback (for testing): just print the path.
        print(f"Saved to: {tmp.name}\n")


# ======================================================================
#  OPTION 2 — Generate report
# ======================================================================

def count_statuses(path):
    """
    Return dict with counts of new / in_progress / faulty / complete,
    plus 'total' = every real data row (has TableName + StandardTagName),
    excluding header rows, blank rows, and border rows.
    """
    counts = {"new": 0, "in_progress": 0, "faulty": 0, "complete": 0, "total": 0}
    for row in read_rows(path):
        counts["total"] += 1          # read_rows already excludes headers/blanks/borders
        status = row["status"].lower()
        if status == "":
            if is_new_tag(row):
                counts["new"] += 1
        elif status == STATUS_IN_PROGRESS:
            counts["in_progress"] += 1
        elif status == STATUS_FAULTY:
            counts["faulty"] += 1
        elif status == STATUS_COMPLETE:
            counts["complete"] += 1
    return counts


def build_report(folder):
    """
    Scan every .xlsx in the folder, map to a mill code, and return a dict
    mill_code -> counts. Mills with no file are left at zero.
    """
    zero = {"new": 0, "in_progress": 0, "faulty": 0, "complete": 0, "total": 0}
    report = {mill: dict(zero) for mill in REPORT_MILLS}
    files = list_workbook_files(folder)
    for path in files:
        if os.path.basename(path).startswith("~$"):
            continue  # skip Excel lock files
        mill = mill_code_from_filename(path)
        counts = count_statuses(path)
        report[mill] = counts
    return report


def debug_status(path, which="in progress"):
    """List every row whose Status matches `which`, with row number and the
    exact stored value (repr reveals hidden spaces/newlines/casing)."""
    which_l = which.lower()
    hits = 0
    print(f"\n--- Rows with status '{which}' in {os.path.basename(path)} ---")
    for row in read_rows(path):
        if row["status"].lower() == which_l:
            hits += 1
            print(f"row {row['row_number']:>5} | status={row['status']!r} | "
                  f"table={row['table']!r} tag={row['tag_name']!r}")
    print(f"Total matching '{which}': {hits}\n")


def debug_all_statuses(path):
    """Tally every distinct Status value actually present (repr shows oddities)."""
    from collections import Counter
    c = Counter()
    for row in read_rows(path):
        c[row["status"]] += 1
    print(f"\n--- Distinct Status values in {os.path.basename(path)} ---")
    for value, n in sorted(c.items(), key=lambda kv: -kv[1]):
        print(f"{n:>5}  {value!r}")
    print()


def print_report(folder):
    report = build_report(folder)
    print("\nDataParc Tag Rollout Report")
    print("=" * 78)
    header = ["Mill", "New", "In Progress", "Faulty", "Completed", "Comments"]
    print("{:<6} {:>4} {:>12} {:>7} {:>14}  {}".format(*header))
    print("-" * 78)
    for mill in REPORT_MILLS:
        c = report.get(mill,
                       {"new": 0, "in_progress": 0, "faulty": 0, "complete": 0, "total": 0})
        completed = "{} of {}".format(c["complete"], c["total"])
        print("{:<6} {:>4} {:>12} {:>7} {:>14}  {}".format(
            mill, c["new"], c["in_progress"], c["faulty"], completed, ""))
    print("=" * 78)
    print()


# ======================================================================
#  Menu
# ======================================================================

def main():
    folder = os.environ.get("WORKBOOK_DIR", ".")
    while True:
        print("Workbook Manager")
        print("  1. Submit to Tag Request Form")
        print("  2. Generate report")
        print("  3. Debug: show status values in a file")
        print("  4. Reconcile with SQL -> Excel output")
        print("  q. Quit")
        choice = input("Choose: ").strip().lower()

        if choice == "1":
            files = list_workbook_files(folder)
            if not files:
                print(f"No workbooks found in {folder!r}.")
                continue
            for i, f in enumerate(files):
                print(f"  [{i}] {os.path.basename(f)}")
            sel = input("File number (or 'all'): ").strip().lower()
            if sel == "all":
                for f in files:
                    print_submit_table(f)
            elif sel.isdigit() and int(sel) < len(files):
                print_submit_table(files[int(sel)])
            else:
                print("Invalid selection.")

        elif choice == "2":
            print_report(folder)

        elif choice == "3":
            files = list_workbook_files(folder)
            if not files:
                print(f"No workbooks found in {folder!r}.")
                continue
            for i, f in enumerate(files):
                print(f"  [{i}] {os.path.basename(f)}")
            sel = input("File number: ").strip()
            if sel.isdigit() and int(sel) < len(files):
                debug_all_statuses(files[int(sel)])
                debug_status(files[int(sel)], "in progress")
            else:
                print("Invalid selection.")

        elif choice == "4":
            import sql_reconcile as sr
            out_path = os.path.join(folder, "reconcile_output.xlsx")
            try:
                conn = _get_sql_connection()
            except Exception as e:
                print(f"Could not connect to SQL: {e}")
                continue
            try:
                path, n_flags, mill_order = sr.build_reconcile_workbook(
                    folder, FILENAME_TO_MILL, conn, out_path)
            finally:
                conn.close()
            print(f"\nWrote {path}")
            print(f"Columns (left to right): "
                  + ", ".join(f"{chr(65+i)}={m}" for i, m in enumerate(mill_order)))
            print(f"{n_flags} discrepancy flag(s) — see the 'Flags' sheet.")
            print("Copy a mill's column and Paste into A1 of that workbook.\n")
            try:
                os.startfile(path)
            except AttributeError:
                pass

        elif choice == "q":
            break
        else:
            print("Unknown option.")


if __name__ == "__main__":
    main()